from django.shortcuts import render, get_object_or_404
from django.forms import modelform_factory
from django.http import HttpResponse
from django.template import loader
from django.shortcuts import redirect
from openpyxl.workbook import Workbook

from empleados.forms import EmpleadoFormulario
from empleados.models import Empleado

# Create your views here.
def agregar_empleado(request):
    pagina = loader.get_template('agregar_empleado.html')
    if request.method == 'GET':
        formulario = EmpleadoFormulario
    elif request.method == 'POST':
        formulario = EmpleadoFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def ver_empleado(request, idEmpleado):
    pagina = loader.get_template('ver_empleado.html')
    empleado = get_object_or_404(Empleado, pk=idEmpleado)
    mensaje = {'empleado': empleado}
    return HttpResponse(pagina.render(mensaje, request))

def editar_empleado(request, idEmpleado):
    pagina = loader.get_template('editar_empleado.html')
    empleado = get_object_or_404(Empleado, pk=idEmpleado)
    if request.method == 'GET':
        formulario = EmpleadoFormulario(instance=empleado)
    elif request.method == 'POST':
        formulario = EmpleadoFormulario(request.POST, instance=empleado)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    mensaje = {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje, request))

def eliminar_empleado(request, idEmpleado):
    empleado = get_object_or_404(Empleado, pk=idEmpleado)
    if empleado:
        empleado.delete()
        return redirect('inicio')

def generar_reporte(request):
    # Obtenemos todas las personas de nuestra base de datos
    #personas = Persona.objects.all()
    empleados = Empleado.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE EMPLEADOS'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:J1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'SEXO'
    ws['E3'] = 'EMAIL'
    ws['F3'] = 'DIRECCIÓN'
    ws['G3'] = 'SUELDO'
    ws['H3'] = 'DEPARTAMENTO'
    ws['I3'] = 'PROYECTO'
    ws['J3'] = 'JORNADA'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for e in empleados:
        ws.cell(row=cont, column=2).value = e.nombre
        ws.cell(row=cont, column=3).value = e.apellido
        ws.cell(row=cont, column=4).value = e.sexo
        ws.cell(row=cont, column=5).value = e.email
        ws.cell(row=cont, column=6).value = e.direccion
        ws.cell(row=cont, column=7).value = e.sueldo
        ws.cell(row=cont, column=8).value = e.departamento.nombre
        ws.cell(row=cont, column=9).value = e.proyecto.nombre
        if e.jornada.tipo == 'M':
            e.jornada.tipo = 'Matutino'
            ws.cell(row=cont, column=10).value = e.jornada.tipo
        elif e.jornada.tipo == 'V':
            e.jornada.tipo = 'Vespertino'
            ws.cell(row=cont, column=10).value = e.jornada.tipo
        elif e.jornada.tipo == 'N':
            e.jornada.tipo = 'Nocturno'
            ws.cell(row=cont, column=10).value = e.jornada.tipo
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteEmpleadosExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response

